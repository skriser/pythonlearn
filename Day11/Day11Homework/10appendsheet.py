#!usr/bin/env python  
# -*- coding:utf-8 -*-
""" 
@time: 2018/05/10 8:55
@author: 柴顺进
@file: 10appendsheet.py 
@software:rongda
@note: 10、在当前工作簿中追加工作表，并获取该工作簿中某一个工作表中的数据添加在新的工作表中
"""  
import openpyxl
from openpyxl import load_workbook

workbook1 = load_workbook('test.xlsx'),
wk_names = workbook1.sheetnames
ws0 = workbook1[wk_names[0]]
ws1 = workbook1.create_sheet('newAJ')

for i, row in enumerate(ws0.iter_rows()):
    for j, cell in enumerate(row):
        ws1.cell(row=i+1, column=j+1, value=cell.value)

workbook1.save('test.xlsx')