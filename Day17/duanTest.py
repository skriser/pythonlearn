#!usr/bin/env python  
# -*- coding:utf-8 -*-
""" 
@time: 2018/05/16 17:38
@author: 柴顺进
@file: duanTest.py 
@software:rongda
@note: 
"""
# -- 4.此题接第3题题干，在第三提的基础上完成以下需求
# -- 编写python代码连接mysql数据库，并向movieRank表中新添加两条数据（数据已给定）
# --
# -- 编写python代码，查询出所有的电影数据，并输出到一个Excel表movieRank.xlsx中,如下图所示
import pymysql
import openpyxl


def getConnect(host, user, password, db, charset):
    return pymysql.connect(host=host, user=user, password=password, db=db, charset=charset)
conn1 = getConnect('localhost', 'root', 'root', 'movies','utf8')
cur = conn1.cursor()
def insertData(data,conn1):
    sql = "insert into movieRank VALUES('%s','%f','%f','%d','%f')"
    cur.execute(sql % data)
    conn1.commit()
    print("插入{}条数据成功！".format(cur.rowcount))
    cur.close()
    conn1.close()

data = ('犬之岛',617.35,0.0908,2,1309.09)
data1 = ('湮灭',135.34,0.0199,9,5556.77)
insertData(data,conn1)
insertData(data1,conn1)
#查询所有数据
taotal_sql = "select * from movieRank"
cur.execute(taotal_sql)
reulst = cur.fetchall()
dict_data = []
biaotou = cur.description
for i in biaotou:
    dict_data.append(i[0])

dict_value = []
for j in reulst:
    dict_value.append(j)

wb = openpyxl.Workbook()
mysheetnames = wb.sheetnames
ws = wb[mysheetnames[0]]

ws.append(dict_data)
print(dict_data)
print(dict_value)
for n in range(len(dict_data)):
    # ws.append(list(dict_data[n]))
    a = (dict_data[n])
#     print(a)
for m in range(len(dict_value)):
    ws.append(dict_value[m])

wb.save('homework.xlsx')